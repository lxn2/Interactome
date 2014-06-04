'''
This script is used to upload the abstracts to s3 in JSON, and users and abstract links to dynamodb.
    Note: Some characters will come out as mojibake (seemingly random stuff...). It's extremely difficult to fix from here.

This script will most likely require changing for using with different spreadsheets of data. 
The WRITES_ALLOWED_PER_LOOP constant in the functions will need to be changed based on dynamodb tables' throughputs.

Requirements for this script:

The abstracts excel file MUST have these as its columns (in order, but spelling doesn't matter)
ControlNumber  PRESENTATIONNUMBER  ABSTRACT TITLE  Abstract    SUBCLASS    CATEGORY-SUBCAT-SUBCLASS    KEYWORD1    KEYWORD2    KEYWORD3    KEYWORD4    SESSIONTITLE    PRESENTERFIRSTNAME  PRESENTERLASTNAME   PRESENTERINSTITUTIONNAME    PRESENTERCONTACTEMAIL

The users excel file MUST have these as its columns (in order, but spelling doesn't matter)
ACTIVITYNAME    PROGRAMSTATUSNAME   CONTROLNUMBER   PRESENTATIONNUMBER  AUTHORFIRSTNAME AUTHORLASTNAME  INSTITUTIONNAME CONTACTEMAIL

Abstracts and Users excel files must correspond to eachother and not to data already in the database. 
I.E., a newly uploaded user will not be attached to an abstract if it was uploaded prior to running the script.


    Note: this has only been tested on xlsx files, but will most likely work with any excel extention.
'''

import sys
import logging
import csv
import json 
import argparse
import time
import pysolr
from boto.s3.key import Key
from boto.dynamodb2.table import Table
from boto import s3
from boto import dynamodb2
from openpyxl import load_workbook

SEQUENCER_TABLE_NAME = 'Sequencer'
USER_TABLE_NAME = 'User'
PAPER_TABLE_NAME = 'Paper'
ABSTRACT_BUCKET_NAME = 'sagebionetworks-interactome-abstracts'
PRESENTATION_NUM_EXCEL_FIELD = 'PresentationNumber'
LASTNAME_EXCEL_FIELD = 'LastName'
FIRSTNAME_EXCEL_FIELD = 'FirstName'
# These are defined by the credentials csv downloaded from amazon
SECRET_KEY_EXCEL_FIELD = 'Secret Access Key'
ACCESS_KEY_EXCEL_FIELD = 'Access Key Id'
# Tells how to format the logging and what file to store it in
logging.basicConfig(filename='dataLoader.log', level=logging.INFO, format='%(asctime)s -- %(levelname)s: %(message)s')

# connect to solr
solr = pysolr.Solr('http://ec2-54-201-190-162.us-west-2.compute.amazonaws.com:8983/solr/', timeout=10)

#solr.delete(q='*:*')

''' 
    Parses the csv file for ACCESS_KEY_EXCEL_FIELD and SECRET_KEY_EXCEL_FIELD.
    It will then use those credentials to connect to s3, dynamo, and all tables.
    This does not check if the keys have enough authority to add to these resources (this may be a good addition in the future).
'''
def connectToAWS(authCSVFileName):
    REGION = 'us-west-2'
    # These resources are used throughout the application.
    global s3Conn
    global dynamoConn
    global sequencerTable
    global usersTable
    global papersTable
    global abstractToPaperDict
    
    try:
    # Opens the credentials and then uses the first line. Using DictReader may be overkill
        csvfile = open(authCSVFileName, 'rU') 
        reader = csv.DictReader(csvfile) 
        authLine = reader.next()
        accessKey = authLine[ACCESS_KEY_EXCEL_FIELD]
        secretKey = authLine[SECRET_KEY_EXCEL_FIELD]
        # Get S3 and dynamo connection
        s3Conn = s3.connect_to_region(REGION, aws_access_key_id=accessKey, aws_secret_access_key=secretKey)
        dynamoConn = dynamodb2.connect_to_region(REGION, aws_access_key_id=accessKey, aws_secret_access_key=secretKey)
        
        sequencerTable = Table(SEQUENCER_TABLE_NAME, connection=dynamoConn)
        usersTable = Table(USER_TABLE_NAME, connection=dynamoConn)
        papersTable = Table(PAPER_TABLE_NAME, connection=dynamoConn)
        abstractToPaperDict = {}
    except Exception, e:
        print(e)
        sys.exit(1)

'''
    Grabs the sequence from the sequence table. If it fails, it will return None. Will return a string of the number on success.
        Note: This should retry if failing for an access reason. If it fails for some other weird reason it will return None.
'''
def getSequence():
    sequenceNumber = None
    try:
      sequenceNumber = dynamoConn.update_item(SEQUENCER_TABLE_NAME, 
                                    {"Id":{"N":"1"}}, 
                                    attribute_updates={"Sequence":{"Action":"ADD", "Value":{"N":"1"}}}, 
                                    expected=None, 
                                    return_values="UPDATED_NEW", 
                                    return_consumed_capacity=None, 
                                    return_item_collection_metrics=None)['Attributes']['Sequence']['N']
      sequenceNumber = str(sequenceNumber)
    except Exception, e:
        print(e)
    return sequenceNumber

'''
    Parses the excel file and adds the users found within to the users table with their connection to an abstract.
    This will edit the papersTable to add an entry into their Authors attribute.
    Will check for duplicate users based off of email address. 
'''
def addUsers(excelFileName):
    # 1 write per loop for both the user and papers tables of dynamodb
    WRITES_ALLOWED_PER_LOOP = 32
    # Load the excel stuff. See http://pythonhosted.org/openpyxl/api.html#module-openpyxl-reader-iter-worksheet-optimized-reader
    workbook = load_workbook(filename = excelFileName, use_iterators = True)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    rowIndex = 0

    for row in worksheet.iter_rows():
        rowIndex += 1
        # skip row with header info. (Don't use iter to make sure we only read one row at a time into memory)
        if(rowIndex == 1):
            continue
        # put to sleep for a second so that we don't go over the throughput allowed
        elif(rowIndex % WRITES_ALLOWED_PER_LOOP == 0):
            time.sleep(1)
        try:
            ctrlNum = "" if (row[2].internal_value is None) else str(row[2].internal_value)
            presNum = "" if (row[3].internal_value is None) else str(row[3].internal_value)
            hashKey = (ctrlNum + presNum).replace(".0", "")
            firstname = row[4].internal_value
            lastname = row[5].internal_value
            institution = row[6].internal_value
            email = row[7].internal_value
            

            # query to see if user was already added
            userId = ''
            userQueryResults = usersTable.query(Email__eq=email, index='Email-index')
            listUserQueryResults = list(userQueryResults)
            # We check all results due to moving institutions, changing names, and such.
            for user in listUserQueryResults:
                if(user['Institution'] == institution and user['LastName'] == lastname and user['FirstName'] == firstname):
                    userId = user['Id']
                    break
            dynamoPaperId = abstractToPaperDict[hashKey] if (hashKey in abstractToPaperDict) else ""
            try:

                solr.add([
                    {
                    'id': userId,
                    'lastName': lastname,
                    'firstName': firstname,
                    'institution': institution
                    }])

                # If new user (email not in dynamo already)
                if(userId == ''):
                    sequenceNumber = getSequence()
                    if (sequenceNumber == None):
                        logging.error("Unabled to sequence item. Index not added: " + str(rowIndex))
                        continue
                    userId = "User" + sequenceNumber

                    newItem = { 'Id': userId,
                        'LastName': lastname,
                        'FirstName': firstname,
                        'Email': email,
                        'Institution': institution,
                    }
                    if(dynamoPaperId != ""):
                        newItem['Papers'] = set([dynamoPaperId])
                    usersTable.put_item(data=newItem)

                # Else it's an old user and has a paper to add to them
                elif(hashKey in abstractToPaperDict):
                    userItem = usersTable.get_item(Id=userId)
                    if('Papers' in userItem):
                        # Fix to this issue https://github.com/boto/boto/issues/1565
                        papersCopy = userItem['Papers']
                        papersCopy.add(dynamoPaperId)
                        userItem['Papers'] = papersCopy
                    else:
                        userItem['Papers'] = set([dynamoPaperId])
                    userItem.save()

                # Else, trying to add a user already in dynamo without a paper to add to them
                else:
                    logging.error("Unabled to find hashkey in dict. Index not added: " + str(rowIndex))
                    continue
                # Update paper to have author
                if(dynamoPaperId != ""):
                    paperItem = papersTable.get_item(Id=dynamoPaperId)
                    if('Authors' in paperItem):
                        authorCopy = paperItem['Authors']
                        authorCopy += "," + userId
                        paperItem['Authors'] = authorCopy
                    else:
                        paperItem['Authors'] = userId
                    paperItem.save()
                else:
                    logging.error("Unable to find dynamopaperid. Index not added: " + str(rowIndex))
            except Exception, e:
                logging.error("Unable to add to usersTable. Index not added: " + str(rowIndex) + "\nError: " + str(e))
        except Exception, e:
            logging.error("Unknown abstract error. Index not added: " + str(rowIndex) + "\nErr: " + str(e))



'''
    Parses the excel file for abstracts. The abstract and its metadata will be dumped into S3. The URL will be stored in the papers table.
    The control and presentation numbers are used as a key to a dict for use in addUsers function.
        Warning: does not check for dupes
'''
def addAbstracts(excelFileName):
    # Write throughput for dynamodb table, one write per loop
    WRITES_ALLOWED_PER_LOOP = 32

    # Load the excel stuff. See http://pythonhosted.org/openpyxl/api.html#module-openpyxl-reader-iter-worksheet-optimized-reader
    workbook = load_workbook(filename = excelFileName, use_iterators = True)
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)

    abstractsBucket = s3Conn.get_bucket(ABSTRACT_BUCKET_NAME)
    abstractKey = Key(abstractsBucket)
    rowIndex = 0

    for row in worksheet.iter_rows():
        rowIndex += 1
        # skip row with header info. (Don't use iter to make sure we only read one row at a time into memory)
        if(rowIndex == 1):
            continue
        # put to sleep for a second so that we don't go over the throughput allowed
        elif(rowIndex % WRITES_ALLOWED_PER_LOOP == 0):
            time.sleep(1)
        try:
            #control number + presentation number
            ctrlNum = "" if (row[0].internal_value is None) else str(row[0].internal_value)
            presNum = "" if (row[1].internal_value is None) else str(row[1].internal_value)
            # Bad conversions sometimes add a .0 to the values
            hashKey = (ctrlNum + presNum).replace(".0", "")

            s3Format = None
            abstractTitle = ""
            try:
                abstractTitle = row[2].internal_value.encode('utf-8')

                # If the cell is empty we replace it with a blank string.
                # The abstract and its title are not checked because we do not want 
                #   anything that doesn't have both of those.
                # Presenter stuff probably doesn't need to be encoded but is anyway.
                subclass = "" if(row[4].internal_value is None) else row[4].internal_value.encode('utf-8')
                catSubcatSubclass = "" if(row[5].internal_value is None) else row[5].internal_value.encode('utf-8')
                keyword1 = "" if (row[6].internal_value is None) else row[6].internal_value.encode('utf-8')
                keyword2 = "" if (row[7].internal_value is None) else row[7].internal_value.encode('utf-8')
                keyword3 = "" if (row[8].internal_value is None) else row[8].internal_value.encode('utf-8')
                keyword4 = "" if (row[9].internal_value is None) else row[9].internal_value.encode('utf-8')
                sessionTitle = "" if (row[10].internal_value is None) else row[10].internal_value.encode('utf-8')
                presenterFirstname = "" if (row[11].internal_value is None) else row[11].internal_value.encode('utf-8')
                presenterLastname = "" if (row[12].internal_value is None) else row[12].internal_value.encode('utf-8')
                presenterInstitution = "" if (row[13].internal_value is None) else row[13].internal_value.encode('utf-8')
                presenterEmail = "" if (row[14].internal_value is None) else row[14].internal_value.encode('utf-8')

                s3Format = {"AbstractTitle": abstractTitle,
                    "Abstract": row[3].internal_value.encode('utf-8'),
                    "Subclass": subclass,
                    "Category-Subcat-Subclass": catSubcatSubclass,
                    "Keyword1": keyword1,
                    "Keyword2": keyword2,
                    "Keyword3": keyword3,
                    "Keyword4": keyword4,
                    "SessionTitle": sessionTitle,
                    "PresenterFirstname": presenterFirstname,
                    "PresenterLastname": presenterLastname,
                    "PresenterInstitution": presenterInstitution,
                    "PresenterEmail": presenterEmail }
            except:
                 logging.error("S3Formatting failed. Index not added: " + str(rowIndex))
                 continue

            sequenceNumber = getSequence()
            if (sequenceNumber == None):
                logging.error("Unabled to sequence item. Index not added: " + str(rowIndex))
                continue

            abstractKey.key = 'Abstract' + sequenceNumber +'.json'
            abstractKey.set_metadata("Content-Type", 'application/json')
            rowJSON = json.dumps(s3Format, skipkeys=False, ensure_ascii=False, sort_keys=True)
            abstractKey.set_contents_from_string(rowJSON)
            abstractKey.make_public()
            abstractUrlLink = abstractKey.generate_url(0, query_auth=False, force_http=True)

            # Now link the S3 url and abstract title to an entry in the dynamodb table
            dynamoPaperId = 'Paper'+ sequenceNumber
            newItem = {'Id': dynamoPaperId, 'Link': abstractUrlLink, 'Title': abstractTitle}
            try:
                papersTable.put_item(data=newItem)

                #Add the solr document into the solr index
                solr.add([
                    {
                        'id': dynamoPaperId,
                        'title': abstractTitle,
                        'text': row[3].internal_value.encode('utf-8'),
                        'keyword1': keyword1,
                        'keyword2': keyword2,
                        'keyword3': keyword3,
                        'keyword4': keyword4
                    }])

            except:
                logging.error("Unabled to add to papers table. Index not added: " + str(rowIndex))
            else:
                # This is how we link Authors to their papers and vice versa.
                abstractToPaperDict[hashKey] = dynamoPaperId

        except Exception, e:
            logging.error("Unknown abstract error. Index not added: " + str(rowIndex) + "\nErr: " + str(e))


def main(argv=sys.argv):

    parser = argparse.ArgumentParser(description='Process optional download flags')

    AUTH_HELP_TEXT = "path to credential csv file from Amazon. This will not be stored. CSV fieldnames must be \"" \
                        + SECRET_KEY_EXCEL_FIELD + "\" for secret key and \"" + ACCESS_KEY_EXCEL_FIELD +"\" for access key"
    parser.add_argument('--auth', dest='usersAuth', help=AUTH_HELP_TEXT, required=True)
    parser.add_argument('--users', dest='usersPath', help='path to users csv file', default=argparse.SUPPRESS, required=True)
    parser.add_argument('--abstracts', dest='abstractsPath', help='path to abstracts csv file', default=argparse.SUPPRESS, required=True)
    options = parser.parse_args(argv[1:])

    connectToAWS(options.usersAuth) # Sets up all resources

    # Abstracts MUST come before users.
    addAbstracts(options.abstractsPath)
    addUsers(options.usersPath)

if __name__ == "__main__":
    sys.exit(main())
